from openpyxl import load_workbook

wb = load_workbook(filename='readexcel.xlsx')
ws = wb.active
print(ws['A2'].value)
print('---------------')
for i in range(1,5):
    for j in range(1,5):
        print(ws.cell(i,j).value) # If you don't add '.value' it also includes cell position
print('---------------')
