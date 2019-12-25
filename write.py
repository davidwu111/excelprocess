from openpyxl import Workbook

wb = Workbook()

ws =wb.active
content = 'string to write'

for i in range(1,5):
    ws.cell(i,1).value = content

print(ws.cell(1,1).value) # Able to read cell value as well.
wb.save('writeexcel.xlsx')
